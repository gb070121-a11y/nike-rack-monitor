from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import asyncio
import base64
import io
import json
import os
from typing import List
from datetime import datetime

from database import (
    init_db, save_rack_scan, get_store_overview,
    get_recent_changes, get_rack_history, search_sku,
    delete_rack, get_excel_data, get_client,
    RACK_MAP_GIMHAE, RACK_MAP_JEONGGWAN, get_rack_number
)
from analyzer import analyze_images_batch

app = FastAPI(title="Nike Rack Monitor V2")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()
app.mount("/static", StaticFiles(directory="static"), name="static")


# ── 페이지 라우트 ──
@app.get("/", response_class=HTMLResponse)
async def root():
    with open("static/index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/floorplan", response_class=HTMLResponse)
async def floorplan_page():
    with open("static/floorplan.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page():
    with open("static/dashboard.html", "r", encoding="utf-8") as f:
        return f.read()


# ── 이미지 읽기 헬퍼 ──
async def read_images(files: List[UploadFile]) -> list:
    images = []
    for f in files:
        data = await f.read()
        b64 = base64.b64encode(data).decode("utf-8")
        images.append({
            "filename": f.filename,
            "b64": b64,
            "content_type": f.content_type or "image/jpeg"
        })
    return images


def get_store_rack_map(store: str) -> dict:
    """매장별 랙맵 반환"""
    return RACK_MAP_JEONGGWAN if store == "jeonggwan" else RACK_MAP_GIMHAE


# ============================================================
# 스캔 API
# ============================================================

@app.post("/api/scan/rack")
async def scan_rack(
    store: str = Form(...),
    rack_name: str = Form(...),
    files: List[UploadFile] = File(...)
):
    """랙별 스캔 - 평면도에서 클릭"""
    rack_map = get_store_rack_map(store)
    if rack_name not in rack_map:
        raise HTTPException(status_code=400, detail=f"알 수 없는 랙: {rack_name} (매장: {store})")

    images = await read_images(files)
    rack_number = rack_map[rack_name]
    racks = await analyze_images_batch(images, start_rack_num=rack_number, single_rack=True)
    products = racks[0]["products"] if racks else []
    result = save_rack_scan(store, rack_name, products)
    return JSONResponse(result)


@app.post("/api/scan/bulk")
async def scan_bulk(
    store: str = Form(...),
    files: List[UploadFile] = File(...),
    rack_names: List[str] = Form(...)
):
    """한번에 스캔 - 스캔 페이지에서 사용
    rack_names: 랙 이름 리스트 (파일 순서와 동일)
    스캔 페이지에서는 숫자(RACK 1, 2...)를 보내므로 숫자면 순서 기반으로 처리
    """
    images = await read_images(files)
    if len(rack_names) != len(images):
        raise HTTPException(status_code=400, detail="파일 수와 랙 이름 수가 다릅니다")

    rack_map = get_store_rack_map(store)

    # rack_name별로 그룹핑
    rack_image_map: dict = {}
    for img, rname in zip(images, rack_names):
        rack_image_map.setdefault(rname, []).append(img)

    async def analyze_one(rname: str, imgs: list):
        # 숫자 문자열이면 rack_map에서 찾고, 없으면 순서 번호로 사용
        rack_number = rack_map.get(rname) or int(rname) if rname.isdigit() else 99
        racks = await analyze_images_batch(imgs, start_rack_num=rack_number, single_rack=True)
        products = racks[0]["products"] if racks else []
        return save_rack_scan(store, rname, products)

    results = await asyncio.gather(*[
        analyze_one(rname, imgs) for rname, imgs in rack_image_map.items()
    ])

    return JSONResponse({
        "store": store,
        "rack_count": len(results),
        "total_products": sum(r["products_count"] for r in results),
        "changed_racks": sum(1 for r in results if r["changes"]["has_changes"]),
        "racks": results,
    })


# ============================================================
# 조회 API
# ============================================================

@app.get("/api/overview/{store}")
async def overview(store: str):
    return JSONResponse(get_store_overview(store))

@app.get("/api/changes/{store}")
async def changes(store: str, limit: int = 50):
    return JSONResponse(get_recent_changes(store, limit))

@app.get("/api/rack/{store}/{rack_name}/history")
async def rack_history_api(store: str, rack_name: str):
    return JSONResponse(get_rack_history(store, rack_name))

# !! /api/search/all 이 /api/search/{store} 보다 먼저 등록되어야 함
@app.get("/api/search/all")
async def search_all_stores(sku: str):
    """김해 + 정관 동시 SKU 검색"""
    gimhae = search_sku("gimhae", sku)
    jeonggwan = search_sku("jeonggwan", sku)
    return JSONResponse({
        "sku": sku,
        "gimhae": gimhae,
        "jeonggwan": jeonggwan,
        "total": len(gimhae) + len(jeonggwan)
    })

@app.get("/api/search/{store}")
async def search(store: str, sku: str):
    return JSONResponse(search_sku(store, sku))


# ============================================================
# 관리 API
# ============================================================

@app.delete("/api/rack/{store}/{rack_name}")
async def delete_rack_api(store: str, rack_name: str):
    delete_rack(store, rack_name)
    return JSONResponse({"success": True})


# ============================================================
# 복사 API
# ============================================================

@app.post("/api/copy/rack")
async def copy_rack(
    from_store: str = Form(...),
    from_rack: str = Form(...),
    to_store: str = Form(...),
    to_rack: str = Form(...),
):
    """랙 전체 데이터를 다른 매장/랙으로 복사"""
    db = get_client()
    src = db.table("rack_master")\
        .select("products")\
        .eq("store", from_store)\
        .eq("rack_name", from_rack)\
        .execute().data

    if not src:
        raise HTTPException(status_code=404, detail=f"{from_store} / {from_rack} 데이터 없음")

    products = src[0]["products"]
    result = save_rack_scan(to_store, to_rack, products)
    return JSONResponse({
        "success": True,
        "from": f"{from_store} / {from_rack}",
        "to": f"{to_store} / {to_rack}",
        "products_count": len(products),
        "changes": result["changes"],
    })


@app.post("/api/copy/products")
async def copy_products(
    skus: List[str] = Form(...),
    from_store: str = Form(...),
    to_store: str = Form(...),
    to_rack: str = Form(...),
):
    """선택한 품번들을 다른 매장/랙으로 복사 (기존 데이터에 추가)"""
    db = get_client()

    # from_store 전체에서 해당 SKU 찾기
    all_racks = db.table("rack_master")\
        .select("products")\
        .eq("store", from_store)\
        .execute().data

    sku_set = set(skus)
    found_products = [
        p for rack in all_racks
        for p in rack["products"]
        if p.get("sku") in sku_set
    ]

    if not found_products:
        raise HTTPException(status_code=404, detail="해당 품번을 찾을 수 없습니다")

    # to_rack 기존 데이터
    existing = db.table("rack_master")\
        .select("products")\
        .eq("store", to_store)\
        .eq("rack_name", to_rack)\
        .execute().data
    existing_products = existing[0]["products"] if existing else []

    # 중복 제거 후 합치기
    existing_skus = {p.get("sku") for p in existing_products}
    new_products = [p for p in found_products if p.get("sku") not in existing_skus]
    merged = existing_products + new_products

    result = save_rack_scan(to_store, to_rack, merged)
    return JSONResponse({
        "success": True,
        "copied_count": len(new_products),
        "skipped_count": len(found_products) - len(new_products),
        "to": f"{to_store} / {to_rack}",
        "total_products": len(merged),
    })


# ============================================================
# 수동 입력 API
# ============================================================

@app.post("/api/manual/rack")
async def manual_input_rack(
    store: str = Form(...),
    rack_name: str = Form(...),
    products_json: str = Form(...),
    mode: str = Form(default="replace"),
):
    """수동으로 제품 정보 입력 (replace: 전체 교체, append: 추가)"""
    try:
        new_products = json.loads(products_json)
    except Exception:
        raise HTTPException(status_code=400, detail="products_json 파싱 실패")

    def normalize_sku(sku: str) -> str:
        sku = sku.strip().upper()
        # 공백을 - 로 변환: AA1234 111 → AA1234-111
        sku = re.sub(r'^([A-Z]{2}[0-9]{4})\s+([0-9]{3})$', r'-', sku)
        sku = re.sub(r'^([0-9]{6})\s+([0-9]{3})$', r'-', sku)
        return sku

    cleaned = []
    for p in new_products:
        sku = normalize_sku(str(p.get("sku", "")))
        if not sku:
            continue
        price = p.get("price")
        sale_price = p.get("sale_price")
        discount_rate = p.get("discount_rate")

        # 할인율로 할인가 자동 계산
        if price and discount_rate and not sale_price:
            try:
                sale_price = round(int(price) * (1 - int(discount_rate) / 100))
            except Exception:
                pass

        # 할인가로 할인율 자동 계산
        if price and sale_price and not discount_rate:
            try:
                discount_rate = round((1 - int(sale_price) / int(price)) * 100)
            except Exception:
                pass

        cleaned.append({
            "sku": sku,
            "name": str(p.get("name", "")).strip(),
            "price": int(price) if price else None,
            "sale_price": int(sale_price) if sale_price else None,
            "discount_rate": int(discount_rate) if discount_rate else None,
            "source": "manual",
        })

    if mode == "append":
        db = get_client()
        existing = db.table("rack_master")\
            .select("products")\
            .eq("store", store)\
            .eq("rack_name", rack_name)\
            .execute().data
        existing_products = existing[0]["products"] if existing else []
        new_skus = {p["sku"] for p in cleaned}
        merged = [p for p in existing_products if p.get("sku") not in new_skus] + cleaned
        result = save_rack_scan(store, rack_name, merged)
    else:
        result = save_rack_scan(store, rack_name, cleaned)

    return JSONResponse({
        "success": True,
        "rack_name": rack_name,
        "products_saved": len(cleaned),
        "changes": result["changes"],
        "mode": mode,
    })


# ============================================================
# 비교 분석 API
# ============================================================

def build_sku_map(overview: dict) -> dict:
    sku_map = {}
    for rack in overview["racks"]:
        for p in rack["products"]:
            sku = p.get("sku", "").strip()
            if not sku:
                continue
            sku_map.setdefault(sku, []).append({
                "rack_name": rack["rack_name"],
                "price": p.get("price"),
                "sale_price": p.get("sale_price"),
                "discount_rate": p.get("discount_rate"),
                "name": p.get("name", ""),
            })
    return sku_map


@app.get("/api/analysis/overlap")
async def overlap_analysis():
    gimhae_map = build_sku_map(get_store_overview("gimhae"))
    jeonggwan_map = build_sku_map(get_store_overview("jeonggwan"))
    all_skus = set(gimhae_map.keys()) | set(jeonggwan_map.keys())

    gimhae_only, jeonggwan_only, both = [], [], []

    for sku in sorted(all_skus):
        in_g = sku in gimhae_map
        in_j = sku in jeonggwan_map
        entry = {
            "sku": sku,
            "gimhae": gimhae_map.get(sku, []),
            "jeonggwan": jeonggwan_map.get(sku, []),
        }
        if in_g and in_j:
            g_price = gimhae_map[sku][0].get("sale_price") or gimhae_map[sku][0].get("price") or 0
            j_price = jeonggwan_map[sku][0].get("sale_price") or jeonggwan_map[sku][0].get("price") or 0
            try:
                entry["price_diff"] = int(j_price) - int(g_price)
            except Exception:
                entry["price_diff"] = 0
            both.append(entry)
        elif in_g:
            gimhae_only.append(entry)
        else:
            jeonggwan_only.append(entry)

    return JSONResponse({
        "summary": {
            "gimhae_only": len(gimhae_only),
            "jeonggwan_only": len(jeonggwan_only),
            "both": len(both),
            "total_unique": len(all_skus),
        },
        "gimhae_only": gimhae_only,
        "jeonggwan_only": jeonggwan_only,
        "both": both,
    })


@app.get("/api/analysis/overlap/excel")
async def overlap_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    g_map = build_sku_map(get_store_overview("gimhae"))
    j_map = build_sku_map(get_store_overview("jeonggwan"))
    all_skus = sorted(set(g_map.keys()) | set(j_map.keys()))

    wb = openpyxl.Workbook()

    def hdr(ws, cols, color):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "전체통합"
    hdr(ws1, ["품번","구분","김해_랙","김해_정가","김해_할인가","정관_랙","정관_정가","정관_할인가"], "FF4D00")
    for sku in all_skus:
        in_g, in_j = sku in g_map, sku in j_map
        status = "양쪽" if (in_g and in_j) else ("김해만" if in_g else "정관만")
        g = g_map[sku][0] if in_g else {}
        j = j_map[sku][0] if in_j else {}
        ws1.append([sku, status,
            g.get("rack_name",""), g.get("price",""), g.get("sale_price",""),
            j.get("rack_name",""), j.get("price",""), j.get("sale_price","")])

    ws2 = wb.create_sheet("양쪽공통")
    hdr(ws2, ["품번","김해_랙","김해_정가","김해_할인가","정관_랙","정관_정가","정관_할인가","가격차"], "39D353")
    for sku in all_skus:
        if sku in g_map and sku in j_map:
            g, j = g_map[sku][0], j_map[sku][0]
            gp = g.get("sale_price") or g.get("price") or 0
            jp = j.get("sale_price") or j.get("price") or 0
            try: diff = int(jp) - int(gp)
            except: diff = ""
            ws2.append([sku,
                g.get("rack_name",""), g.get("price",""), g.get("sale_price",""),
                j.get("rack_name",""), j.get("price",""), j.get("sale_price",""), diff])

    ws3 = wb.create_sheet("김해만")
    hdr(ws3, ["품번","랙","정가","할인가","할인율"], "58A6FF")
    for sku in all_skus:
        if sku in g_map and sku not in j_map:
            g = g_map[sku][0]
            ws3.append([sku, g.get("rack_name",""), g.get("price",""), g.get("sale_price",""), g.get("discount_rate","")])

    ws4 = wb.create_sheet("정관만")
    hdr(ws4, ["품번","랙","정가","할인가","할인율"], "E3B341")
    for sku in all_skus:
        if sku in j_map and sku not in g_map:
            j = j_map[sku][0]
            ws4.append([sku, j.get("rack_name",""), j.get("price",""), j.get("sale_price",""), j.get("discount_rate","")])

    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(12, w + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"overlap_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# 엑셀 다운로드
# ============================================================

@app.get("/api/excel/{store}")
async def download_excel(store: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    data = get_excel_data(store)
    overview = data["overview"]
    changes_data = data["changes"]

    wb = openpyxl.Workbook()

    def style_header(ws, headers, color):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "전체현황"
    style_header(ws1, ["랙이름","랙번호","품번","정가","할인가","할인율","최근스캔"], "FF4D00")
    for rack in overview["racks"]:
        for p in rack["products"]:
            ws1.append([
                rack["rack_name"], rack["rack_number"],
                p.get("sku",""), p.get("price",""), p.get("sale_price",""),
                f"{p['discount_rate']}%" if p.get("discount_rate") else "",
                rack["last_scanned_at"]
            ])

    ws2 = wb.create_sheet("변동이력")
    style_header(ws2, ["스캔일시","랙이름","신규","삭제","변동","제품수"], "E3B341")
    for h in changes_data:
        ch = h.get("changes", {})
        s = ch.get("summary", {})
        ws2.append([h["scanned_at"], h["rack_name"],
            s.get("added",0), s.get("removed",0), s.get("changed",0), h["product_count"]])

    for ws in [ws1, ws2]:
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(12, w + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rack_{store}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
